from tkinter import*
import sqlite3
from tkinter import ttk
from openpyxl.workbook import Workbook #Create workbook in our own
from openpyxl import load_workbook # file that has been already created


def register_user():

	user_info = username.get()
	pw_info = password.get()

	conn = sqlite3.connect('Form.db')
	with conn:
		cursur = conn.cursor()
	cursur.execute("CREATE TABLE IF NOT EXISTS user(username TEXT NOT NULL, password TEXT NOT NULL);")
	cursur.execute('INSERT INTO user(username, password) VALUES(?,?)',(user_info,pw_info))
	conn.commit()
	Label(w2, text = "Registration Success", fg = "green", font = ("calibri",  11)).pack()

def  login_verify():

	while True:
		user1 = user_verify.get()
		pw1 = pw_verify.get()
		with sqlite3.connect("Form.db") as db:
			cursur = db.cursor()
		find_user = ("SELECT * FROM user WHERE username = ? AND password = ?")
		cursur.execute(find_user,[(user1), (pw1)])
		result = cursur.fetchall()
		if result:
			f_window = Tk()


		# Frame design
			wrapper = LabelFrame(f_window, text="Product Class Information")
			wrapper.pack(padx=10, pady=5, fill="both", expand="yes")
			wrapper2 = LabelFrame(f_window, text="Tire Wheel Information")
			wrapper2.pack(padx=10, pady=10, fill="both", expand="yes")

			#levels for Wrapper
			Label1 = Label(wrapper, text = "Product Class")
			Label2 = Label(wrapper, text = "Front Installation/ride height")
			Label10 = Label(wrapper, text = "Variant Combination")
			Label11 = Label(wrapper, text = "Actual Height")
			Label3 = Label(wrapper2, text = "Manufacture")
			Label4 = Label(wrapper2, text = "Dimensions")
			Label5 = Label(wrapper2, text = "Tire Position")
			Label6 = Label(wrapper2, text = "Static Loaded Radius (in)")
			Label7 = Label(wrapper2, text = "Static Loaded Radius (mm)")
			Label8 = Label(wrapper2, text = "Unloaded Diameter (in)")
			Label9 = Label(wrapper2, text = "Unloaded Diameter (mm)")

			Label1.grid(row=1, column=0, sticky='w')
			Label2.grid(row=2, column=0, sticky='w')
			Label10.grid(row=3, column=0, sticky='w')
			Label11.grid(row=4, column=0, sticky='w')
			Label3.grid(row=1, column=0, sticky='w')
			Label4.grid(row=2, column=0, sticky='w')
			Label5.grid(row=3, column=0, sticky='w')
			Label6.grid(row=4, column=0, sticky='w')
			Label7.grid(row=5, column=0, sticky='w')
			Label8.grid(row=6, column=0, sticky='w')
			Label9.grid(row=7, column=0, sticky='w')

			# Product Class
			var1 = StringVar()
			VC = ttk.Combobox(wrapper, textvariable = var1, width=15)
			VC["values"] = [" ",
			'VS-LHAUL,FST-PAR,FSC5.4',
			'VS-LHAUL,FST-PAR,FSC5.7',
			'VS-LHAUL,FST-PAR,FSC6.0',
			'VS-LHAUL,FST-PAR,FSC6.6',
			'VS-LHAUL,FST-PAR7,FSC5.4',
			'VS-LHAUL,FST-PAR7,FSC5.7',
			'VS-LHAUL,FST-PAR7,FSC6.0',
			'VS-LHAUL,FST-PAR7,FSC6.6',
			'VS-RHAUL,FST-PAR,FSC5.4',
			'VS-RHAUL,FST-PAR,FSC5.7',
			'VS-RHAUL,FST-PAR,FSC6.0',
			'VS-RHAUL,FST-PAR,FSC6.6',
			'VS-RHAUL,FST-PAR7,FSC5.4',
			'VS-RHAUL,FST-PAR7,FSC5.7',
			'VS-RHAUL,FST-PAR7,FSC6.0',
			'VS-RHAUL,FST-PAR7,FSC6.6',]
			VC.grid(row=3, column=3, padx=10, pady=10)
			VC.current(0) 						

			# Front Installation
			var2 = StringVar()
			FI = ttk.Combobox(wrapper, textvariable = var2, width=15)
			FI["values"] = [" ","FIH185","FIH220",]
			FI.grid(row=2, column=3, padx=10, pady=10)
			FI.current(0) 
			# Front Installation
			var4 = StringVar()
			AH = ttk.Combobox(wrapper, textvariable = var4, width=15)
			AH["values"] = [" ","185","220","215","250"]
			AH.grid(row=4, column=3, padx=10, pady=10)
			AH.current(0) 

			# Variant Combination
			var3 = StringVar()
			PC = ttk.Combobox(wrapper, textvariable = var3, width=15)
			PC["values"] = [" ","B1, B2",]
			PC.grid(row=1, column=3, padx=10, pady=10)
			PC.current(0) 

			# Manufacturing
			var = StringVar()
			Manufacturing = ttk.Combobox(wrapper2, textvariable = var, width=15)
			Manufacturing["values"] = [" ","Michelin", "Goodyear", "Bridgestone", "Continental",
				"Yokohama"]
			Manufacturing.grid(row=1, column=3, padx=10, pady=10)
			Manufacturing.current(0) # set to default

			# Drop boxes for Wrapper2
			Dimensions = ttk.Combobox(wrapper2, value = [" "], width=15)
			Dimensions.grid(row=2, column = 3, padx=10, pady=10)
			Dimensions.current(0)

			# Drop boxes for Wrapper2
			var = StringVar()
			TP = ttk.Combobox(wrapper2, value = [" "], textvariable = var, width=15)
			TP.grid(row=3, column = 3, padx=10, pady=10)
			TP.current(0)

			# Drop down list for Static Loaded Radius (in)
			SLR_in = ttk.Combobox(wrapper2, value = [" "], width=15)
			SLR_in.grid(row=4, column = 3, padx=10, pady=10)
			SLR_in.current(0)

			# Drop down list for Static Loaded Radius (in)
			SLR_mm = ttk.Combobox(wrapper2, value = [" "], width=15)
			SLR_mm.grid(row=5, column = 3, padx=10, pady=10)
			SLR_mm.current(0)

			# Drop down list for Static Loaded Radius (in)
			UND_in = ttk.Combobox(wrapper2, value = [" "], width=15)
			UND_in.grid(row=6, column = 3, padx=10, pady=10)
			UND_in.current(0)

			# Drop down list for Static Loaded Radius (in)
			UND_mm = ttk.Combobox(wrapper2, value = [" "], width=15)
			UND_mm.grid(row=7, column = 3, padx=10, pady=10)
			UND_mm.current(0)


			# Getting dimentions for each manufacturing from excel 
			wb = Workbook()
			wb = load_workbook("Dim_mich.xlsx")
			ws = wb.active

			column_a = ws['A']
			column_b = ws['B']
			column_c = ws['C']
			column_d = ws['D']
			column_e = ws['E']

			list = ''
			list1 = ''
			list2 = ''
			list3 = ''
			list4 = ''

			for cell in column_a:
				list = f'{list + str(cell.value)}\n'
			for cell in column_b:
				if cell.value is None:
					continue
				else:
					list1 = f'{list1 + str(cell.value)}\n'
			for cell in column_c:
				if cell.value is None:
					continue
				else:
					list2 = f'{list2 + str(cell.value)}\n'
			for cell in column_d:
				if cell.value is None:
					continue
				else:
					list3 = f'{list3 + str(cell.value)}\n'
			for cell in column_e:
				if cell.value is None:
					continue
				else:
					list4 = f'{list4 + str(cell.value)}\n'	


			Mich_dim = list
			gy_dim = list1
			Bridg_dim = list2
			Cont_dim = list3
			Yoko_dim = list4


			def pick_man(event):

				if Manufacturing.get() == "Michelin":
					Dimensions.config(value=Mich_dim)
					if(Dimensions.get() == '10R22.5'):
							TP.config(value=[' ','All'])
							SLR_in.config(value = [ ' ','18.7'])
							SLR_mm.config(value=[' ','474.98'])
							UND_in.config(value=[' ','40.1'])
							UND_mm.config(value=[' ','1018.54'])

					elif(Dimensions.get() == '11R22.5'):
						TP.config(value=[' ','All', 'Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.1','19.2','19.3','19.6'])
							SLR_mm.config(value=[ ' ','485.14','487.65','490.22','497.84'])
							if SLR_in.get() == '19.1':
								SLR_mm.config(value=['485.14'])
								UND_in.config(value=[' ','41.3'])
								UND_mm.config(value=[' ','1049.02'])
							else:
								if SLR_mm.get() == '485.14':
									SLR_mm.config(value=['19.1'])
									UND_in.config(value=[' ','41.3'])
									UND_mm.config(value=[' ','1049.02'])

							if SLR_in.get() == '19.2':
								SLR_mm.config(value=['487.65'])
							else:
								if SLR_mm.get() == '487.65':
									SLR_in.config(value=['19.2'])

							if SLR_in.get() == '19.3':
								SLR_mm.config(value=['490.22'])
							else:
								if SLR_mm.get() == '490.22':
									SLR_in.config(value=['19.3'])

							if SLR_in.get() == '19.6':
								SLR_mm.config(value=['497.84'])
							else:
								if SLR_mm.get() == '497.84':
									SLR_in.config(value=['19.6'])

						if(TP.get() == 'Drive'):
							SLR_in.config(value = [ ' ','19.6'])
							SLR_mm.config(value=[ ' ','497.84'])
					elif(Dimensions.get() == '11R24.5'):
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','20.2','20.3','20.5'])
							SLR_mm.config(value=[ ' ','513.08','515.62','520.7'])
						if(TP.get() == 'Drive'):
							SLR_in.config(value = [ ' ','20.4','20.5','20.6','20.7','20.8'])
							SLR_mm.config(value=[ ' ','518.16','520.7','523.24','525.78','528.32'])
					elif(Dimensions.get() == '12R22.5'):
						TP.config(value=[' ','All','Diver'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.8','20','20.1'])
							SLR_mm.config(value=[ ' ','502.92','508','510.54'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','19.9','20'])
							SLR_mm.config(value=[ ' ','505.46','508',])
					elif(Dimensions.get() == '12R24.5'):
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','21'])
							SLR_mm.config(value=[ ' ','533.4'])
							UND_in.config(value=[' ','44.9'])
							UND_mm.config(value=[' ','1140.46'])
					elif(Dimensions.get() == '245/70R19.5'):
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','15.6'])
							SLR_mm.config(value=[ ' ','396.24'])
							UND_in.config(value=[' ','33.6'])
							UND_mm.config(value=[' ','853.44'])
					elif(Dimensions.get() == '255/70R22.5'):
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','17.2'])
							SLR_mm.config(value=[ ' ','436.88'])
							UND_in.config(value=[' ','36.7'])
							UND_mm.config(value=[' ','932.18'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','17.4'])
							SLR_mm.config(value=[ ' ','441.96'])
							UND_in.config(value=[' ','37.2'])
							UND_mm.config(value=[' ','944.88'])
					elif(Dimensions.get() == '255/80R22.5'):
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','17.9'])
							SLR_mm.config(value=[ ' ','454.66'])
							UND_in.config(value=[' ','38.5'])
							UND_mm.config(value=[' ','977.9'])
					elif(Dimensions.get() == '265/70R19.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','15.8'])
							SLR_mm.config(value=[ ' ','401.32'])
							UND_in.config(value=[' ','34.3'])
							UND_mm.config(value=[' ','871.22'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','15.8','15.9'])
							SLR_mm.config(value=[ ' ','401.32','403.86'])
							UND_in.config(value=[' ','34.2','34.4'])
							UND_mm.config(value=[' ','868.68','873.76'])
					elif(Dimensions.get() == '275/70R22.5'):
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','17.6'])
							SLR_mm.config(value=[ ' ','447.04'])
							UND_in.config(value=[' ','37.8','38'])
							UND_mm.config(value=[' ','960.12','965.2'])
					elif(Dimensions.get() == '275/80R22.5'): #check with sponsor
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','18.6','18.7','18.8'])
							SLR_mm.config(value=[ ' ','472.44','474.98','477.52'])
						if TP.get() == "Drive": 
							SLR_in.config(value = [ ' ','18.7','18.8','19.9','19'])
							SLR_mm.config(value=[ ' ','474.98','477.52','480.06','482.6'])
					elif(Dimensions.get() == '275/80R24.5'):
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.3'])
							SLR_mm.config(value=[ ' ','490.22'])
							UND_in.config(value=[' ','41.3']) 
							UND_mm.config(value=[' ','1049.02'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','19.5','19.6'])
							SLR_mm.config(value=[ ' ','495.3','497.84'])

					elif(Dimensions.get() == '285/70R19.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','16.2'])
							SLR_mm.config(value=[ ' ','411.48'])
							UND_in.config(value=[' ','35.2'])
							UND_mm.config(value=[' ','894.08'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','16.3'])
							SLR_mm.config(value=[ ' ','414.02'])
							UND_in.config(value=[' ','35.4'])
							UND_mm.config(value=[' ','899.16'])
					elif(Dimensions.get() == '295/60R22.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','16.7'])
							SLR_mm.config(value=[ ' ','424.18'])
							UND_in.config(value=[' ','36.1'])
							UND_mm.config(value=[' ','916.94'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','16.9'])
							SLR_mm.config(value=[ ' ','429.26'])
							UND_in.config(value=[' ','36.5'])
							UND_mm.config(value=[' ','916.94'])
					elif(Dimensions.get() == '295/80R22.5'): 
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.1','19.3'])
							SLR_mm.config(value=[ ' ','485.14','490.22'])
							UND_in.config(value=[' ','41.3','41.5'])
							UND_mm.config(value=[' ','1049.02','1054.1'])
					elif(Dimensions.get() == '305/75R24.5'): 
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' '])
							SLR_mm.config(value=[ ' '])
							UND_in.config(value=[' '])
							UND_mm.config(value=[' '])
					elif(Dimensions.get() == '315/80R22.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.5','19.6','19.7','19.8'])
							SLR_mm.config(value=[ ' ','495.3','497.84','500.38','502.92'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','20'])
							SLR_mm.config(value=[ ' ','508'])
					elif(Dimensions.get() == '365/70R22.5'): 
						TP.config(value=[' ','Steer'])
						if TP.get() == "Steer":
							SLR_in.config(value = [ ' ','19.6'])
							SLR_mm.config(value=[ ' ','497.84'])
							UND_in.config(value=[' ','42.5'])
							UND_mm.config(value=[' ','1079.5'])
							
					elif(Dimensions.get() == '385/65R22.5'): 
						TP.config(value=[' ','All','Steer'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.6'])
							SLR_mm.config(value=[ ' ','497.84'])
						if TP.get() == "Steer":
							SLR_in.config(value = [ ' ','19.6'])
							SLR_mm.config(value=[ ' ','497.84'])

					elif(Dimensions.get() == '425/65R22.5'): 
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','20.6'])
							SLR_mm.config(value=[ ' ','523.24'])
							UND_in.config(value=[' ','44.5','44.7','44.8'])
							UND_mm.config(value=[' ','1130.3','1135.38','1137.92'])
					elif(Dimensions.get() == '445/50R22.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','18.4'])
							SLR_mm.config(value=[ ' ','467.36'])
							UND_in.config(value=[' ','39.7'])
							UND_mm.config(value=[' ','1008.38'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','18.3','18.5','18.6','18.7'])
							SLR_mm.config(value=[ ' ','464.82','469.9','472.44','474.98'])
					elif(Dimensions.get() == '445/65R22.5'): 
						TP.config(value=[' ','All'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','21','21.1','21.2'])
							SLR_mm.config(value=[ ' ','533.4','535.94','538.48'])
							UND_in.config(value=[' ','41.3','41.5'])
							UND_mm.config(value=[' ','1158.24','1163.32','1168.4'])
					elif(Dimensions.get() == '455/55R22.5'): 
						TP.config(value=[' ','All','Drive'])
						if TP.get() == "All":
							SLR_in.config(value = [ ' ','19.3','19.4','19.5'])
							SLR_mm.config(value=[ ' ','490.22','492.76','495.3'])
							UND_in.config(value=[' ','41.6','41.9','42.2'])
							UND_mm.config(value=[' ','1056.64','1064.26','1071.88'])
						if TP.get() == "Drive":
							SLR_in.config(value = [ ' ','19.4','19.6'])
							SLR_mm.config(value=[ ' ','492.76','497.84'])
							UND_in.config(value=[' ','42','42.4'])
							UND_mm.config(value=[' ','1066.8','1076.96'])
					else:
						clear()
				if Manufacturing.get() == "Bridgestone":
					Dimensions.config(value=Bridg_dim)
					if(Dimensions.get() == '11R22.5'):
							TP.config(value=[' ','All','Drive',])
							SLR_in.config(value = [ ' ','18.7'])
							SLR_mm.config(value=[' ','474.98'])
							UND_in.config(value=[' ','40.1'])
							UND_mm.config(value=[' ','1018.54'])
				if Manufacturing.get() == "Goodyear":
					Dimensions.config(value=gy_dim)
				if Manufacturing.get() == "Continental":
					Dimensions.config(value=Cont_dim)
				if Manufacturing.get() == "Yokohama":
					Dimensions.config(value=Yoko_dim)


			# Clear combox but not values
			def clear1():

				PC.set('') 
				PC.config(value=[' ',])
				VC.set('') 
				VC.config(value = [ ' ',])
				FI.set('') 
				FI.config(value=[' ',])
				AH.set('') 
				AH.config(value=[' ',])
				

			# Clear combox but not values
			def clear():
				TP.set('') 
				TP.config(value=[' ',])
				SLR_in.set('') 
				SLR_in.config(value = [ ' ',])
				SLR_mm.set('') 
				SLR_mm.config(value=[' ',])
				UND_in.set('') 
				UND_in.config(value=[' ',])
				UND_mm.set('') 
				UND_mm.config(value=[' ',])


			Manufacturing.bind("<<ComboboxSelected>>", pick_man)
			Dimensions.bind("<<ComboboxSelected>>", pick_man)
			TP.bind("<<ComboboxSelected>>", pick_man)
			SLR_in.bind("<<ComboboxSelected>>", pick_man)
			SLR_mm.bind("<<ComboboxSelected>>", pick_man)

			rst_button = Button(wrapper, text="Clear Data", command=clear1)
			rst_button.grid(row=5, column=3, padx=10, pady=10)

			rst_button = Button(wrapper2, text="Clear Data", command=clear)
			rst_button.grid(row=8, column=3, padx=10, pady=10)

			sel_button = Button(wrapper, text="Select Data",)
			sel_button.grid(row=5, column=1, padx=10, pady=10)

			sel_button = Button(wrapper2, text="Select Data",)
			sel_button.grid(row=8, column=1, padx=10, pady=10)	
								

			
		
			f_window.geometry("400x600")
			f_window.title("Volvo's Trucks Data")
			f_window.mainloop()
			break	
		else:
			print("No recognized")
			break




def register():
	global w2
	w2 = Toplevel(w1)
	w2.title("Register")
	w2.geometry("300x250")

	global username
	global password
	global user_entry
	global pw_entry
	username = StringVar()
	password = StringVar()


	Label(w2, text = "Please enter details below ").pack()
	Label(w2, text = "").pack()
	Label(w2, text = "Username Email ").pack()
	user_entry = Entry(w2, textvariable = username)
	user_entry.pack()
	Label(w2, text = "Password ").pack()
	pw_entry = Entry(w2, textvariable = password)
	pw_entry.pack()
	Label(w2, text = "").pack()
	Button(w2, text = "Register", width = 10, height = 1, command = register_user).pack()
def login():
	global w3
	w3 = Toplevel(w1)
	w3.title("Login")
	w3.geometry("300x250")
	Label(w3, text = "Please enter details below ").pack()
	Label(w3, text = "").pack()

	global user_verify
	global pw_verify
	user_verify = StringVar()
	pw_verify = StringVar()

	global user_entry1
	global pw_entry1

	Label(w3, text = "Username Email ").pack()
	user_entry1 = Entry(w3, textvariable = user_verify)
	user_entry1.pack()
	Label(w3, text = "Password ").pack()
	pw_entry1 = Entry(w3, textvariable = pw_verify, show = "*")
	pw_entry1.pack()
	Label(w3, text = "").pack()
	Button(w3, text = "Login", width = 10, height = 1, command = login_verify).pack()

	




def main_window():
	global w1
	w1 = Tk()
	w1.geometry("400x600")
	w1.title("Volvo")
	Label(text ="Volvo's Trucks Home Page", bg = "grey", width ="300", height = "2", font = ("Calibri, 13") ).pack()
	Label(text = " ").pack()
	Button(text = "Login", width ="30", height = "2", command = login).pack()
	Label(text = " ").pack()
	Button(text = "Register", width ="30", height = "2", command = register).pack()

	w1.mainloop()
main_window()

